import yfinance as yf
import numpy as np
import pandas as pd
from pathlib import Path
from nicpy.CacheDict import CacheDict
from nicpy.nic_webscrape import get_check_country_code
import openpyxl
from datetime import datetime



def get_last_prices(tickers):
    # TODO: yf_tickers for non-ASX holdings?
    yf_tickers = [el + '.AX' for el in tickers]
    last_prices = {}
    data = yf.download(tickers = yf_tickers, period = "1d", interval = "1m", group_by = 'ticker')
    for ticker, yf_ticker in zip(tickers, yf_tickers):
        closes = data[yf_ticker]['Close']
        remove_nans = closes[~np.isnan(closes)]
        last_price = remove_nans.iloc[-1]
        last_price_time = remove_nans.index[-1]
        last_prices[ticker] = (last_price_time, last_price)
    return last_prices


def update_ETF_spreadsheet_prices():
    # Open the first sheet
    file_path = r'F:\Nick\Misc\Finances\ETF performance and portfolio\ETF performance and portfolio.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[wb.sheetnames[0]]

    # Find the ETF tickers being tracked
    tickers, ticker, i = [], '', 0
    while ticker != 'TOTALS':
        ticker = sheet.cell(row=5+i, column=2).value
        if ticker != 'TOTALS':
            tickers.append(ticker)
            i += 1

    # Get latest prices for those tickers
    last_prices = get_last_prices(tickers)

    # Put the prices into the spreadsheet and update the today date
    for i, ticker in enumerate(tickers):
        sheet.cell(row=5+i, column=4).value = last_prices[ticker][1]
    sheet.cell(row=4, column=21).value = datetime.now().strftime('%m/%d/%Y')
    wb.save(file_path)

def read_format_ETF_holdings(csv_directory):
    csv_directory = Path(csv_directory)
    holdings = {}

    country_code_cache = CacheDict(get_check_country_code, persist_filename = 'read_ETF_holdings.dat', persist_lifetime_hours=24)
    code_edits = {'UM': 'US'}

    # Vanguard - 'VAE', 'VAP', 'VAS', 'VEU', 'VGS', 'VTS'
    for ticker in ['VAE', 'VAP', 'VAS', 'VEU', 'VGS', 'VTS']:
        print(ticker)
        holdings[ticker] = pd.read_csv(str(csv_directory / '{}.csv'.format(ticker)), header=3)
        keep_cols = ['Ticker', 'Sector', 'Country code', '% of net assets']
        holdings[ticker] = holdings[ticker][keep_cols]
        holdings[ticker].dropna(how='any', subset=['% of net assets', 'Sector', 'Country code'], inplace=True)
        holdings[ticker].rename(columns={'% of net assets': 'Weight (%)'}, inplace=True)
        holdings[ticker] = holdings[ticker].applymap(lambda x: format_vanguard(x))
        holdings[ticker] = holdings[ticker][holdings[ticker]['Country code'] != '—']
        holdings[ticker]['Country code'] = holdings[ticker]['Country code'].apply(lambda x: country_code_cache.get_key_value((x, code_edits))[0])
        holdings[ticker] = holdings[ticker].astype({'Weight (%)':float})
        holdings[ticker]['Sector'] = holdings[ticker]['Sector'].apply(lambda x: get_general_sector(x))

    # BetaShares - A200
    ticker = 'A200'
    holdings[ticker] = pd.read_csv(str(csv_directory/'{}.csv'.format(ticker)), header=6)
    holdings[ticker].dropna(how='any', subset=['Weight (%)', 'Sector', 'Country'], inplace=True)
    keep_cols = ['Ticker', 'Sector', 'Country', 'Weight (%)']
    holdings[ticker] = holdings[ticker][keep_cols]
    holdings[ticker]['Weight (%)'] = holdings[ticker]['Weight (%)']*100
    holdings[ticker].rename(columns={'Country': 'Country code'}, inplace=True)
    holdings[ticker]['Country code'] = holdings[ticker]['Country code'].apply(lambda x: country_code_cache.get_key_value((x, code_edits))[0])
    holdings[ticker]['Ticker'] = holdings[ticker]['Ticker'].apply(lambda x: x.split(' ')[0])
    # Convert to iShares sectors
    sector_map = {'Communication Services':'Communication', 'Healthcare': 'Health Care'}
    holdings[ticker]['Sector'] = holdings[ticker]['Sector'].apply(lambda x: sector_map[x] if x in sector_map.keys() else x)

    # iShares - IJR
    ticker = 'IJR'
    holdings[ticker] = pd.read_csv(str(csv_directory / '{}.csv'.format(ticker)), header=25)
    keep_cols = ['Ticker', 'Sector', 'Location', 'Weight (%)']
    holdings[ticker] = holdings[ticker][keep_cols]
    holdings[ticker].rename(columns={'Location': 'Country code'}, inplace=True)
    holdings[ticker].dropna(how='any', subset=['Weight (%)', 'Sector', 'Country code'], inplace=True)
    holdings[ticker]['Country code'] = holdings[ticker]['Country code'].apply(lambda x: country_code_cache.get_key_value((x, code_edits))[0])

    # iShares - IVV
    ticker = 'IVV'
    holdings[ticker] = pd.read_csv(str(csv_directory / '{}.csv'.format(ticker)), header=26)
    keep_cols = ['Ticker', 'Sector', 'Location', 'Weight (%)']
    holdings[ticker] = holdings[ticker][keep_cols]
    holdings[ticker].rename(columns={'Location': 'Country code'}, inplace=True)
    holdings[ticker].dropna(how='any', subset=['Weight (%)', 'Sector', 'Country code'], inplace=True)
    holdings[ticker]['Country code'] = holdings[ticker]['Country code'].apply(lambda x: country_code_cache.get_key_value((x, code_edits))[1])

    return holdings


def format_vanguard(el_str):
    return el_str.replace('"', '').replace('=', '').replace('%', '').replace('$', '')

def get_general_sector(specific_sector):
    general_sectors_with_keywords = {
        'Real Estate'               : ['REIT', 'REITs', 'real estate'],
        'Cash and/or Derivatives'   : ['cash', 'derivative'],
        'Communication'             : ['publishing', 'comms','communication', 'satellite', 'advertising'],
        'Consumer Discretionary'    : ['specialise','specialize','specialty','leisure','restaurants','hardware', 'merchandise', 'tobacco', 'movies', 'entertainment', 'soft drinks', 'luxury', 'gaming', 'brewer', 'consumer electronics', 'department stores'],
        'Consumer Staples'          : ['home', 'personal', 'Food', 'houseware', 'education', 'meat', 'hypermarket', 'supermarket', 'household', 'home improvement', 'footwear', 'apparel'],
        'Energy'                    : ['power','Oil', 'Renewable', 'coal'],
        'Financials'                : ['insurance', 'Financial', 'finance', 'Fintech', 'Bank', 'multi-sector holdings'],
        'Health Care'               : ['Health', 'pharma', 'biotech', 'life sciences', 'drug'],
        'Industrials'               : ['support services','motorcycle','truck', 'distributor', 'rubber', 'consulting', 'paper', 'marine', 'employment', 'Steel', 'electrical', 'highway', 'railtrack', 'cruise', 'hotel', 'forest', 'electronic', 'distiller', 'Automobile', 'agricultur', 'auto parts', 'construction', 'engineer', 'freight', 'farm', 'Automotive', 'printing', 'Semiconductor', 'gold', 'silver', 'industrial', 'building', 'chemical', 'metal', 'copper', 'railroad', 'aerospace', 'defense', 'airlines', 'airport', 'carriers', 'aluminum'],
        'Information Technology'    : ['IT', 'information', 'data', 'software', 'interactive media', 'internet', 'broadcasting'],
        'Utilities'                 : ['utility']}
    for gs in general_sectors_with_keywords.keys():
        if specific_sector.upper() == gs.upper(): return gs
        for keyword in general_sectors_with_keywords[gs]:
            if keyword.upper() in specific_sector.upper(): return gs
    return 'Other'

def export_holdings_sector_country_weights(holdings):
    sector_weights, country_weights = {ticker: {} for ticker in holdings.keys()}, {ticker: {} for ticker in holdings.keys()}
    for ticker in holdings.keys():
        for sector in sorted(set(holdings[ticker]['Sector'])):
            sector_data = holdings[ticker][holdings[ticker]['Sector'] == sector]
            sector_weights[ticker][sector] = sector_data['Weight (%)'].sum()
        for country in sorted(set(holdings[ticker]['Country code'])):
            country_data = holdings[ticker][holdings[ticker]['Country code'] == country]
            country_weights[ticker][country] = country_data['Weight (%)'].sum()
        sector_weights[ticker]['total']  = sum(list(sector_weights[ticker].values()))
        country_weights[ticker]['total'] = sum(list(country_weights[ticker].values()))
        sector_weights_df = pd.DataFrame(sector_weights)
        sector_weights_df.sort_index(inplace=True)
        country_weights_df = pd.DataFrame(country_weights)
        country_weights_df.sort_index(inplace=True)
        country_code_cache = CacheDict(get_check_country_code, persist_filename='read_ETF_holdings.dat', persist_lifetime_hours=24)
        country_weights_df['Country'] = country_weights_df.index
        country_weights_df.index = country_weights_df['Country'].apply(lambda x: x+'/'+country_code_cache.get_key_value((x, {}))[1])
        country_weights_df.drop(columns=['Country'], inplace=True)
        sector_weights_df.to_excel('sector_weights_python.xlsx')
        country_weights_df.to_excel('country_weights_python.xlsx')
    return country_weights_df, country_weights_df



if __name__ == '__main__':

    #################### Get last prices
    # my_ETFs = ['IJR', 'VAP', 'VAS', 'VEU', 'VGE', 'VGS', 'VTS', 'VAE', 'VGB', 'IVV', 'A200']
    # last_prices = get_last_prices(my_ETFs)

    update_ETF_spreadsheet_prices()

    # #################### Load holdings data from providers' CSVs and calculate the sector and country weights
    # csv_directory=Path(r'F:\Nick\Misc\Finances\ETF PDSs')
    # holdings = read_format_ETF_holdings(csv_directory)
    # VAS_set = sorted(set(holdings['VAS']['Sector']))
    # VAP_set = sorted(set(holdings['VAP']['Sector']))
    # VGS_set = sorted(set(holdings['VGS']['Sector']))
    # A200_set = sorted(set(holdings['A200']['Sector']))
    # IJR_set = sorted(set(holdings['IJR']['Sector']))
    # stillthere = [el for el in VGS_set if el not in IJR_set]
    # stillthere2 = [el for el in VAS_set if el not in IJR_set]
    # sector_weights, country_weights = export_holdings_sector_country_weights(holdings)

